from __future__ import annotations

import shutil
import uuid
from pathlib import Path, PureWindowsPath
from typing import Dict, List

import json
import os
import time
import difflib

import requests
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

APP_ROOT = Path(__file__).resolve().parent
IMG_ROOT = Path(os.getenv("IMG_ROOT", r"D:\saloon\Looks\img"))
SESSIONS_ROOT = Path(os.getenv("SESSIONS_ROOT", r"D:\saloon\Looks\sessions"))
STATIC_ROOT = APP_ROOT / "static"
PROMPTS_XLSX = Path(os.getenv("PROMPTS_XLSX", r"D:\saloon\Looks\HairstylePresertPromts.xlsx"))

LIGHTX_API_KEY = os.getenv("LIGHTX_API_KEY")
LIGHTX_UPLOAD_URL = "https://api.lightxeditor.com/external/api/v2/uploadImageUrl"
LIGHTX_HAIRSTYLE_URL = "https://api.lightxeditor.com/external/api/v2/hairstyle"
LIGHTX_STATUS_URL = "https://api.lightxeditor.com/external/api/v2/order-status"

API_PROVIDER = os.getenv("HAIRSTYLE_PROVIDER", "lightx").strip().lower()
API_MARKET_KEY = os.getenv("API_MARKET_KEY")
API_MARKET_UPLOAD_URL = "https://prod.api.market/api/v1/magicapi/image-upload/upload"
API_MARKET_HAIRSTYLE_URL = "https://prod.api.market/api/v1/lightxeditor/hairstyle/hairstyle/"
API_MARKET_STATUS_URL = "https://prod.api.market/api/v1/lightxeditor/hairstyle/order-status/"

app = FastAPI(title="LOOKS Local MVP")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class Preset:
    def __init__(self, preset_id: str, category: str, major: str, filename: str, path: Path):
        self.preset_id = preset_id
        self.category = category
        self.major = major
        self.filename = filename
        self.path = path

    def to_dict(self) -> Dict[str, str]:
        return {
            "id": self.preset_id,
            "category": self.category,
            "major": self.major,
            "filename": self.filename,
            "displayId": IDS_BY_FILENAME.get(self.filename.lower(), ""),
            "imageUrl": f"/img/{self.category}/{self.major}/{self.filename}",
        }


def build_preset_id(category: str, major: str, filename: str) -> str:
    return f"{category}__{major}__{filename}"


def scan_presets() -> Dict[str, Preset]:
    presets: Dict[str, Preset] = {}
    if not IMG_ROOT.exists():
        return presets
    for category_dir in IMG_ROOT.iterdir():
        if not category_dir.is_dir():
            continue
        category = category_dir.name
        for major_dir in category_dir.iterdir():
            if not major_dir.is_dir():
                continue
            major = major_dir.name
            for file_path in major_dir.iterdir():
                if file_path.suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
                    continue
                preset_id = build_preset_id(category, major, file_path.name)
                presets[preset_id] = Preset(preset_id, category, major, file_path.name, file_path)
    return presets


PRESETS: Dict[str, Preset] = scan_presets()
PROMPTS_BY_FILENAME: Dict[str, str] = {}
IDS_BY_FILENAME: Dict[str, str] = {}
RESULTS_LIMIT = 5
API_CALLS_LOG = "api_calls.json"


@app.on_event("startup")
def ensure_dirs() -> None:
    SESSIONS_ROOT.mkdir(parents=True, exist_ok=True)
    load_prompts()


def load_prompts() -> None:
    PROMPTS_BY_FILENAME.clear()
    IDS_BY_FILENAME.clear()
    if not PROMPTS_XLSX.exists():
        print(f"[prompts] file not found: {PROMPTS_XLSX}")
        return
    import openpyxl

    wb = openpyxl.load_workbook(PROMPTS_XLSX)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if not headers:
            continue
        header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}
        if "Prompt" not in header_map or "ImagePath" not in header_map:
            continue
        id_idx = header_map.get("ID")
        prompt_idx = header_map["Prompt"]
        path_idx = header_map["ImagePath"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_id = row[id_idx] if id_idx is not None and id_idx < len(row) else None
            prompt = row[prompt_idx] if prompt_idx < len(row) else None
            image_path = row[path_idx] if path_idx < len(row) else None
            if not prompt or not image_path:
                continue
            image_str = str(image_path).strip()
            if "\\" in image_str and "/" not in image_str:
                filename = PureWindowsPath(image_str).name.lower()
            else:
                filename = Path(image_str).name.lower()
            PROMPTS_BY_FILENAME[filename] = str(prompt).strip()
            if row_id is not None:
                if isinstance(row_id, float) and row_id.is_integer():
                    row_id = int(row_id)
                IDS_BY_FILENAME[filename] = str(row_id).strip()
    print(f"[prompts] loaded {len(PROMPTS_BY_FILENAME)} prompts from {PROMPTS_XLSX}")


@app.get("/api/prompts/status")
def prompts_status() -> Dict[str, int]:
    return {
        "loaded_prompts": len(PROMPTS_BY_FILENAME),
        "known_presets": len(PRESETS),
    }


@app.get("/api/prompts/lookup")
def prompts_lookup(filename: str) -> Dict[str, object]:
    key = Path(filename).name.lower()
    exists = key in PROMPTS_BY_FILENAME
    matches = difflib.get_close_matches(key, PROMPTS_BY_FILENAME.keys(), n=5, cutoff=0.6)
    return {"filename": key, "exists": exists, "matches": matches}


def get_prompt_for_preset(preset: Preset) -> str | None:
    return PROMPTS_BY_FILENAME.get(preset.filename.lower())


def get_latest_photo(session_dir: Path) -> Path | None:
    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        candidate = session_dir / f"latest{ext}"
        if candidate.exists():
            return candidate
    return None


def load_results(session_dir: Path) -> List[Dict[str, str]]:
    results_path = session_dir / "results.json"
    if not results_path.exists():
        return []
    try:
        return json.loads(results_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []


def save_results(session_dir: Path, results: List[Dict[str, str]]) -> None:
    results_path = session_dir / "results.json"
    results_path.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")


def load_api_calls(session_dir: Path) -> List[Dict[str, object]]:
    log_path = session_dir / API_CALLS_LOG
    if not log_path.exists():
        return []
    try:
        return json.loads(log_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []


def save_api_calls(session_dir: Path, calls: List[Dict[str, object]]) -> None:
    log_path = session_dir / API_CALLS_LOG
    log_path.write_text(json.dumps(calls, ensure_ascii=False, indent=2), encoding="utf-8")


def upload_image_to_lightx(image_path: Path) -> str:
    if not LIGHTX_API_KEY:
        raise HTTPException(status_code=500, detail="LIGHTX_API_KEY is not set")
    content_type = "image/jpeg" if image_path.suffix.lower() in {".jpg", ".jpeg"} else "image/png"
    size = image_path.stat().st_size
    payload = {"uploadType": "imageUrl", "size": size, "contentType": content_type}
    headers = {"Content-Type": "application/json", "x-api-key": LIGHTX_API_KEY}
    res = requests.post(LIGHTX_UPLOAD_URL, json=payload, headers=headers, timeout=30)
    if res.status_code != 200:
        raise HTTPException(status_code=502, detail=f"LightX uploadImageUrl failed: {res.text}")
    data = res.json()
    print(f"[apimarket] upload response keys: {list(data.keys())}")
    if data.get("statusCode") != 2000:
        raise HTTPException(status_code=502, detail=f"LightX uploadImageUrl error: {data}")
    body = data.get("body") or {}
    upload_url = body.get("uploadImage")
    image_url = body.get("imageUrl")
    if not upload_url or not image_url:
        raise HTTPException(status_code=502, detail="LightX uploadImageUrl missing uploadImage/imageUrl")
    with image_path.open("rb") as f:
        put_res = requests.put(upload_url, data=f, headers={"Content-Type": content_type}, timeout=60)
    if put_res.status_code not in (200, 201):
        raise HTTPException(status_code=502, detail=f"LightX PUT upload failed: {put_res.text}")
    return image_url


def upload_image_to_apimarket(image_path: Path) -> str:
    if not API_MARKET_KEY:
        raise HTTPException(status_code=500, detail="API_MARKET_KEY is not set")
    content_type = "image/jpeg" if image_path.suffix.lower() in {".jpg", ".jpeg"} else "image/png"
    headers = {"accept": "application/json", "x-api-market-key": API_MARKET_KEY}
    with image_path.open("rb") as f:
        files = {"filename": (image_path.name, f, content_type)}
        res = requests.post(API_MARKET_UPLOAD_URL, headers=headers, files=files, timeout=60)
    if res.status_code != 200:
        raise HTTPException(status_code=502, detail=f"API Market upload failed: {res.text}")
    data = res.json()
    root = data.get("root") or data.get("body") or data
    url = root.get("url")
    if not url:
        raise HTTPException(status_code=502, detail=f"API Market upload missing url: {data}")
    return url


def request_hairstyle(image_url: str, prompt: str) -> str:
    if not LIGHTX_API_KEY:
        raise HTTPException(status_code=500, detail="LIGHTX_API_KEY is not set")
    headers = {"Content-Type": "application/json", "x-api-key": LIGHTX_API_KEY}
    res = requests.post(LIGHTX_HAIRSTYLE_URL, json={"imageUrl": image_url, "textPrompt": prompt}, headers=headers, timeout=60)
    if res.status_code != 200:
        raise HTTPException(status_code=502, detail=f"LightX hairstyle failed: {res.text}")
    data = res.json()
    if data.get("statusCode") != 2000:
        raise HTTPException(status_code=502, detail=f"LightX hairstyle error: {data}")
    body = data.get("body") or {}
    order_id = body.get("orderId")
    if not order_id:
        raise HTTPException(status_code=502, detail="LightX hairstyle missing orderId")
    return order_id


def request_hairstyle_apimarket(image_url: str, prompt: str) -> str:
    if not API_MARKET_KEY:
        raise HTTPException(status_code=500, detail="API_MARKET_KEY is not set")
    headers = {"accept": "*/*", "x-api-market-key": API_MARKET_KEY, "Content-Type": "application/json"}
    res = requests.post(API_MARKET_HAIRSTYLE_URL, json={"imageUrl": image_url, "textPrompt": prompt}, headers=headers, timeout=60)
    if res.status_code != 200:
        raise HTTPException(status_code=502, detail=f"API Market hairstyle failed: {res.text}")
    data = res.json()
    if data.get("statusCode") != 2000:
        raise HTTPException(status_code=502, detail=f"API Market hairstyle error: {data}")
    body = data.get("body") or {}
    order_id = body.get("orderId")
    if not order_id:
        raise HTTPException(status_code=502, detail="API Market hairstyle missing orderId")
    return order_id


def poll_hairstyle(order_id: str) -> str:
    headers = {"Content-Type": "application/json", "x-api-key": LIGHTX_API_KEY}
    for _ in range(5):
        time.sleep(3)
        res = requests.post(LIGHTX_STATUS_URL, json={"orderId": order_id}, headers=headers, timeout=30)
        if res.status_code != 200:
            continue
        data = res.json()
        if data.get("statusCode") != 2000:
            continue
        body = data.get("body") or {}
        status = body.get("status")
        if status == "active":
            output = body.get("output")
            if output:
                return output
        if status == "failed":
            break
    raise HTTPException(status_code=502, detail="LightX hairstyle output not ready")


def poll_hairstyle_apimarket(order_id: str) -> str:
    headers = {"accept": "*/*", "x-api-market-key": API_MARKET_KEY, "Content-Type": "application/json"}
    for _ in range(5):
        time.sleep(3)
        res = requests.post(API_MARKET_STATUS_URL, json={"orderId": order_id}, headers=headers, timeout=30)
        if res.status_code != 200:
            continue
        data = res.json()
        if data.get("statusCode") != 2000:
            continue
        body = data.get("body") or {}
        status = body.get("status")
        if status == "active":
            output = body.get("output")
            if output:
                return output
        if status == "failed":
            break
    raise HTTPException(status_code=502, detail="API Market hairstyle output not ready")


@app.get("/api/presets")
def get_presets(category: str | None = None, major: str | None = None) -> List[Dict[str, str]]:
    items = []
    for preset in PRESETS.values():
        if category and preset.category.lower() != category.lower():
            continue
        if major and preset.major.lower() != major.lower():
            continue
        items.append(preset.to_dict())
    return items


@app.get("/api/presets/majors")
def get_majors(category: str) -> List[str]:
    majors = sorted({p.major for p in PRESETS.values() if p.category.lower() == category.lower()})
    return majors


@app.post("/api/session/start")
def start_session(userType: str | None = None) -> Dict[str, str]:
    session_id = uuid.uuid4().hex
    session_dir = SESSIONS_ROOT / session_id
    session_dir.mkdir(parents=True, exist_ok=True)
    if userType:
        meta = {"userType": userType}
        (session_dir / "meta.json").write_text(json.dumps(meta), encoding="utf-8")
    return {"sessionId": session_id}


@app.post("/api/session/upload")
def upload_photo(sessionId: str = Form(...), file: UploadFile = File(...)) -> Dict[str, str]:
    session_dir = SESSIONS_ROOT / sessionId
    if not session_dir.exists():
        raise HTTPException(status_code=404, detail="Session not found")
    ext = Path(file.filename or "").suffix or ".jpg"
    dest = session_dir / f"latest{ext}"
    with dest.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"sessionId": sessionId, "imageUrl": f"/sessions/{sessionId}/{dest.name}"}


@app.post("/api/hairstyle/apply")
def apply_hairstyle(sessionId: str, presetId: str) -> Dict[str, str]:
    session_dir = SESSIONS_ROOT / sessionId
    if not session_dir.exists():
        raise HTTPException(status_code=404, detail="Session not found")
    preset = PRESETS.get(presetId)
    if not preset:
        raise HTTPException(status_code=404, detail="Preset not found")

    results = load_results(session_dir)
    for entry in results:
        if entry.get("presetId") == presetId:
            return {
                "sessionId": sessionId,
                "presetId": presetId,
                "sourcePresetUrl": preset.to_dict()["imageUrl"],
                "resultUrl": entry.get("resultUrl", ""),
            }
    if len(results) >= RESULTS_LIMIT:
        raise HTTPException(status_code=429, detail="Session limit reached")

    prompt = get_prompt_for_preset(preset)
    if not prompt:
        raise HTTPException(status_code=404, detail="Prompt not found for preset")
    source_photo = get_latest_photo(session_dir)
    if not source_photo or not source_photo.exists():
        raise HTTPException(status_code=404, detail="Source photo not found")

    calls = load_api_calls(session_dir)
    call_entry: Dict[str, object] = {
        "presetId": presetId,
        "provider": API_PROVIDER,
        "prompt": prompt,
        "status": "started",
        "startedAt": time.time(),
    }
    calls.append(call_entry)
    save_api_calls(session_dir, calls)

    try:
        if API_PROVIDER == "apimarket":
            image_url = upload_image_to_apimarket(source_photo)
            order_id = request_hairstyle_apimarket(image_url, prompt)
            output_url = poll_hairstyle_apimarket(order_id)
        else:
            image_url = upload_image_to_lightx(source_photo)
            order_id = request_hairstyle(image_url, prompt)
            output_url = poll_hairstyle(order_id)
        call_entry["status"] = "success"
        call_entry["orderId"] = order_id
        call_entry["outputUrl"] = output_url
        call_entry["completedAt"] = time.time()
        save_api_calls(session_dir, calls)
    except Exception as exc:
        call_entry["status"] = "failed"
        call_entry["error"] = str(exc)
        call_entry["completedAt"] = time.time()
        save_api_calls(session_dir, calls)
        raise

    result_dir = session_dir / preset.category
    result_dir.mkdir(parents=True, exist_ok=True)
    result_name = f"result-{presetId.replace('__', '_')}"
    out_path = result_dir / f"{result_name}.jpg"

    if not out_path.exists():
        res = requests.get(output_url, timeout=60)
        if res.status_code != 200:
            raise HTTPException(status_code=502, detail="Failed to download LightX output")
        with out_path.open("wb") as f:
            f.write(res.content)

    result_url = f"/sessions/{sessionId}/{preset.category}/{out_path.name}"
    results.append(
        {
            "presetId": presetId,
            "category": preset.category,
            "filename": out_path.name,
            "resultUrl": result_url,
            "createdAt": time.time(),
        }
    )
    save_results(session_dir, results)

    return {
        "sessionId": sessionId,
        "presetId": presetId,
        "sourcePresetUrl": preset.to_dict()["imageUrl"],
        "resultUrl": result_url,
    }


@app.get("/api/session/{session_id}/results")
def session_results(session_id: str) -> List[Dict[str, str]]:
    session_dir = SESSIONS_ROOT / session_id
    if not session_dir.exists():
        raise HTTPException(status_code=404, detail="Session not found")
    results = []
    for category_dir in session_dir.iterdir():
        if not category_dir.is_dir():
            continue
        for file_path in category_dir.iterdir():
            if file_path.is_file():
                results.append(
                    {
                        "category": category_dir.name,
                        "filename": file_path.name,
                        "imageUrl": f"/sessions/{session_id}/{category_dir.name}/{file_path.name}",
                    }
                )
    return results


app.mount("/img", StaticFiles(directory=IMG_ROOT), name="img")
app.mount("/sessions", StaticFiles(directory=SESSIONS_ROOT), name="sessions")
app.mount("/", StaticFiles(directory=STATIC_ROOT, html=True), name="static")
